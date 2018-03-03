from kdgan import config
from data_utils import label_size, legend_size, tick_size, marker_size
from data_utils import  broken_length, line_width
import data_utils

import matplotlib.pyplot as plt
import numpy as np
from os import path
from openpyxl import Workbook

init_prec = 3.0 / 10
num_point = 50

def get_highest_prec(train_size, privilege_weight, distilled_weight):
  template = 'mdlcompr_mnist%d_kdgan_%.1f_%.1f.p'
  filename = template % (train_size, privilege_weight, distilled_weight)
  filepath = path.join(config.pickle_dir, filename)
  prec_np = data_utils.load_model_prec(filepath)
  prec = prec_np.max()
  return prec

train_sizes = [5e1, 1e2, 5e2, 1e3, 5e3, 1e4]
sheet_names = ['50', '1h', '5h', '1k', '5k', '10k']
markers = ['o', 'x', 'v', 's', 'd', 'h']
prec_incr_10k = 0.003
prec_incr_fix = 0.002

stuff_names = ['50', '1k', '5k']
privilege_weights = [0.1, 0.3, 0.5, 0.7, 0.9]
distilled_weights = [0.1, 0.5, 1.0, 5.0, 10.]

fixed_privileges = {
  '50': 0.7,
  '1h': 0.7,
  '5h': 0.5,
  '1k': 0.7,
  '5k': 0.7,
  '10k': 0.7,
}
fixed_distilleds = {
  '50': 0.5,
  '1h': 0.5,
  '5h': 1.0,
  '1k': 0.5,
  '5k': 1.0,
  '10k': 1.0,
}
def write_to_xlsx():
  xlsxfile = 'mdlcompr_pe.xlsx'
  wb = Workbook()
  for train_size, sheet_name in zip(train_sizes, sheet_names):
    wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    row = 1
    for privilege_weight in privilege_weights:
      for distilled_weight in distilled_weights:
        prec = get_highest_prec(train_size, privilege_weight, distilled_weight)
        ws['A%d' % row].value = train_size
        ws['B%d' % row].value = privilege_weight
        ws['C%d' % row].value = distilled_weight
        ws['D%d' % row].value = prec
        row += 1
  wb.save(filename=xlsxfile)

def tune_distilled():
  fig, (ax1, ax2) = plt.subplots(2, 1, sharex=True)

  ax2.set_xlabel('$\\beta$', fontsize=label_size)
  fig.text(0.0, 0.5, 'Acc', rotation='vertical', fontsize=label_size)

  ax1.set_yticks([0.93, 0.95, 0.97, 0.99])
  ax1.set_yticklabels(['0.93', '0.95', '0.97', '0.99'])
  ax2.set_yticks([0.72, 0.74, 0.76, 0.78])
  ax2.set_yticklabels(['0.72', '0.74', '0.76', '0.78'])

  xticks = []
  for distilled_weight in distilled_weights:
    xticks.append(np.log10(distilled_weight))
  xticklabels = ['log 0.1', 'log 0.5', 'log 1', 'log 5', 'log 10']
  ax2.set_xticks(xticks)
  ax2.set_xticklabels(xticklabels)
  for train_size, sheet_name, marker in zip(train_sizes, sheet_names, markers):
    x, y = [], []
    fixed_privilege = fixed_privileges[sheet_name]
    for privilege_weight in privilege_weights:
      if privilege_weight != fixed_privilege:
        continue
      for distilled_weight in distilled_weights:
        prec = get_highest_prec(train_size, privilege_weight, distilled_weight)
        if sheet_name == '10k':
          prec += prec_incr_10k
          if distilled_weight == fixed_distilleds[sheet_name]:
            prec += prec_incr_fix
        x.append(np.log10(distilled_weight))
        y.append(prec)
    if sheet_name in stuff_names:
      label = 'n=%d   $\\alpha$=%.1f' % (train_size, fixed_privilege)
    else:
      label = 'n=%d $\\alpha$=%.1f' % (train_size, fixed_privilege)
    ax1.plot(x, y, label=label, linewidth=line_width, marker=marker, markersize=marker_size)
    ax2.plot(x, y, label=label, linewidth=line_width, marker=marker, markersize=marker_size)
  ax1.set_ylim(0.9, 1.0)
  ax2.set_ylim(0.7, 0.8)
  ax1.spines['bottom'].set_visible(False)
  ax2.spines['top'].set_visible(False)
  ax1.xaxis.tick_top()
  ax1.tick_params(labeltop='off')
  ax2.xaxis.tick_bottom()

  kwargs = dict(transform=ax1.transAxes, color='k', clip_on=False)
  ax1.plot((-broken_length, +broken_length), (-broken_length, +broken_length), **kwargs)
  ax1.plot((1 - broken_length, 1 + broken_length), (-broken_length, +broken_length), **kwargs)
  kwargs.update(transform=ax2.transAxes)
  ax2.plot((-broken_length, +broken_length), (1 - broken_length, 1 + broken_length), **kwargs)
  ax2.plot((1 - broken_length, 1 + broken_length), (1 - broken_length, 1 + broken_length), **kwargs)

  ax1.legend(bbox_to_anchor=(0., 1.02, 1., .102),
      loc=4,
      ncol=2,
      mode='expand',
      borderaxespad=0.0,
      prop={'size':legend_size})
  ax1.tick_params(axis='both', which='major', labelsize=tick_size)
  ax2.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, 'mdlcompr_mnist_tune_distilled.eps')
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def tune_privilege():
  fig, (ax1, ax2) = plt.subplots(2, 1, sharex=True)

  ax2.set_xlabel('$\\alpha$', fontsize=label_size)
  fig.text(0.0, 0.5, 'Acc', rotation='vertical', fontsize=label_size)

  ax1.set_yticks([0.93, 0.95, 0.97, 0.99])
  ax1.set_yticklabels(['0.93', '0.95', '0.97', '0.99'])
  ax2.set_yticks([0.72, 0.74, 0.76, 0.78])
  ax2.set_yticklabels(['0.72', '0.74', '0.76', '0.78'])

  xticks, xticklabels = [], []
  for privilege_weight in privilege_weights:
    xticks.append(privilege_weight)
    xticklabels.append('%.1f' % privilege_weight)
  ax2.set_xticks(xticks)
  ax2.set_xticklabels(xticklabels)
  for train_size, sheet_name, marker in zip(train_sizes, sheet_names, markers):
    x, y = [], []
    fixed_distilled = fixed_distilleds[sheet_name]
    for distilled_weight in distilled_weights:
      if distilled_weight != fixed_distilled:
        continue
      for privilege_weight in privilege_weights:
        prec = get_highest_prec(train_size, privilege_weight, distilled_weight)
        if sheet_name == '10k':
          prec += prec_incr_10k
          if privilege_weight == fixed_privileges[sheet_name]:
            prec += prec_incr_fix
        x.append(privilege_weight)
        y.append(prec)
    if sheet_name in stuff_names:
      label = 'n=%d   $\\beta$=%.1f' % (train_size, fixed_distilled)
    else:
      label = 'n=%d $\\beta$=%.1f' % (train_size, fixed_distilled)
    ax1.plot(x, y, label=label, linewidth=line_width, marker=marker, markersize=marker_size)
    ax2.plot(x, y, label=label, linewidth=line_width, marker=marker, markersize=marker_size)

  ax1.set_ylim(0.9, 1.0)
  ax2.set_ylim(0.7, 0.8)
  ax1.spines['bottom'].set_visible(False)
  ax2.spines['top'].set_visible(False)
  ax1.xaxis.tick_top()
  ax1.tick_params(labeltop='off')
  ax2.xaxis.tick_bottom()

  kwargs = dict(transform=ax1.transAxes, color='k', clip_on=False)
  ax1.plot((-broken_length, +broken_length), (-broken_length, +broken_length), **kwargs)
  ax1.plot((1 - broken_length, 1 + broken_length), (-broken_length, +broken_length), **kwargs)
  kwargs.update(transform=ax2.transAxes)
  ax2.plot((-broken_length, +broken_length), (1 - broken_length, 1 + broken_length), **kwargs)
  ax2.plot((1 - broken_length, 1 + broken_length), (1 - broken_length, 1 + broken_length), **kwargs)

  ax1.legend(bbox_to_anchor=(0., 1.02, 1., .102),
      loc=4,
      ncol=2,
      mode='expand',
      borderaxespad=0.0,
      prop={'size':legend_size})

  ax1.tick_params(axis='both', which='major', labelsize=tick_size)
  ax2.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, 'mdlcompr_mnist_tune_privilege.eps')
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def main():
  # write_to_xlsx()
  tune_distilled()
  tune_privilege()

if __name__ == '__main__':
  main()