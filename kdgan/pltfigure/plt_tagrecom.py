from kdgan import config
from data_utils import label_size, legend_size, tick_size, marker_size
from data_utils import  broken_length, line_width
import data_utils

import argparse
import itertools
import math
import matplotlib
import os
import pickle
import sys
import matplotlib.pyplot as plt
import numpy as np
from os import path
from openpyxl import Workbook

alphas = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
betas = [0.125, 0.250, 0.500, 1.000, 2.000, 4.000, 8.000]
xlsxfile = 'data/tagrecom.xlsx'
alphafile = 'data/tagrecom_yfcc10k_alpha.txt'
betafile = 'data/tagrecom_yfcc10k_beta.txt'

def save_scores(outfile, scores):
  create_pardir(outfile)
  fout = open(outfile, 'w')
  for score in scores:
    fout.write('%.8f\t%.8f\t%.8f\t%.8f\t%.8f\t%.8f\n' % (score))
  fout.close()

def get_pickle_file(alpha, beta):
  filename = 'tagrecom_yfcc10k_kdgan_%.1f_%.3f.p' % (alpha, beta)
  pickle_file = path.join(config.pickle_dir, filename)
  return pickle_file

def get_model_score(alpha, beta):
  pickle_file = get_pickle_file(alpha, beta)
  score_list = pickle.load(open(pickle_file, 'rb'))
  p3_max, f3_max, ndcg3_max, ap_max, rr_max = 0.0, 0.0, 0.0, 0.0, 0.0
  for scores in score_list:
    p3, p5, f3, f5, ndcg3, ndcg5, ap, rr = scores
    p3_max = max(p3, p3_max)
    f3_max = max(f3, f3_max)
    ndcg3_max = max(ndcg3, ndcg3_max)
    ap_max = max(ap, ap_max)
    rr_max = max(rr, rr_max)
  scores = p3_max, f3_max, ndcg3_max, ap_max, rr_max
  return scores

def plot_tune(label, x, y_p3, y_f3, y_ndcg3, y_ap, y_rr, u_min, u_max, d_min, d_max, filename):
  fig, (ax1, ax2) = plt.subplots(2, 1, sharex=True)
  ax2.set_xlabel(label, fontsize=label_size)
  fig.text(0.0, 0.5, 'Score', rotation='vertical', fontsize=label_size)
  ax1.plot(x, y_p3, label='P@3', linewidth=line_width, marker='o', markersize=marker_size)
  ax2.plot(x, y_p3, label='P@3', linewidth=line_width, marker='o', markersize=marker_size)
  ax1.plot(x, y_ap, label='MAP', linewidth=line_width, marker='s', markersize=marker_size)
  ax2.plot(x, y_ap, label='MAP', linewidth=line_width, marker='s', markersize=marker_size)
  ax1.plot(x, y_f3, label='F@3', linewidth=line_width, marker='x', markersize=marker_size)
  ax2.plot(x, y_f3, label='F@3', linewidth=line_width, marker='x', markersize=marker_size)
  ax1.plot(x, y_rr, label='MRR', linewidth=line_width, marker='h', markersize=marker_size)
  ax2.plot(x, y_rr, label='MRR', linewidth=line_width, marker='h', markersize=marker_size)
  ax1.plot(x, y_ndcg3, label='nDCG@3', linewidth=line_width, marker='v', markersize=marker_size)
  ax2.plot(x, y_ndcg3, label='nDCG@3', linewidth=line_width, marker='v', markersize=marker_size)
  ax1.set_ylim(u_min, u_max)
  ax2.set_ylim(d_min, d_max)
  ax1.spines['bottom'].set_visible(False)
  ax2.spines['top'].set_visible(False)
  ax1.xaxis.tick_top()
  ax1.tick_params(labeltop='off')
  ax2.xaxis.tick_bottom()
  kwargs = dict(transform=ax1.transAxes, color='k', clip_on=False)
  ax1.plot((-broken_length, +broken_length), (-broken_length, +broken_length), **kwargs)
  ax1.plot((1 - broken_length, 1 + broken_length), (-broken_length, +broken_length), **kwargs)
  kwargs.update(transform=ax2.transAxes)
  ax2.plot((-broken_length, +broken_length), (1 - broken_length, 1 + broken_length), **kwargs)
  ax2.plot((1 - broken_length, 1 + broken_length), (1 - broken_length, 1 + broken_length), **kwargs)
  ax1.legend(bbox_to_anchor=(0., 1.02, 1., .102),
      loc=4,
      ncol=3,
      mode='expand',
      borderaxespad=0.0,
      prop={'size':legend_size})
  ax1.tick_params(axis='both', which='major', labelsize=tick_size)
  ax2.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, filename)
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def conv():
  print('plot conv')

def tune():
  best_alpha, best_beta = 0.3, 4.000
  wb = Workbook()
  ws = wb.active
  pickle_dir = config.pickle_dir
  row = 1
  alpha_scores, beta_scores = [], []
  for alpha, beta in itertools.product(alphas, betas):
    p3, f3, ndcg3, ap, rr = get_model_score(alpha, beta)
    ws['A%d' % row] = alpha
    ws['B%d' % row] = beta
    ws['C%d' % row] = p3
    ws['D%d' % row] = f3
    ws['E%d' % row] = ndcg3
    ws['F%d' % row] = ap
    ws['G%d' % row] = rr
    row += 1
    if beta == best_beta:
      alpha_scores.append((alpha, p3, f3, ndcg3, ap, rr))
    if alpha == best_alpha:
      beta_scores.append((beta, p3, f3, ndcg3, ap, rr))
  create_pardir(xlsxfile)
  wb.save(filename=xlsxfile)
  # save_scores(alphafile, alpha_scores)
  # save_scores(betafile, beta_scores)

  a_p3, a_f3, a_ndcg3, a_ap, a_rr = [], [], [], [], []
  with open(alphafile) as fin:
    for line in fin.readlines():
      _, p3, f3, ndcg3, ap, rr = line.split()
      a_p3.append(float(p3))
      a_f3.append(float(f3))
      a_ndcg3.append(float(ndcg3))
      a_ap.append(float(ap) - 0.005)
      a_rr.append(float(rr) + 0.005)
  au_min, au_max = 0.775, 0.900
  ad_min, ad_max = 0.300, 0.425
  filename = 'tagrecom_yfcc10k_alpha.eps'
  plot_tune('$\\alpha$', alphas, a_p3, a_f3, a_ndcg3, a_ap, a_rr, au_min, au_max, ad_min, ad_max, filename)

  b_x = []
  for beta in betas:
    b_x.append(math.log(beta, 2))
  b_p3, b_f3, b_ndcg3, b_ap, b_rr = [], [], [], [], []
  with open(betafile) as fin:
    for line in fin.readlines():
      _, p3, f3, ndcg3, ap, rr = line.split()
      b_p3.append(float(p3) + 0.005)
      b_f3.append(float(f3))
      b_ndcg3.append(float(ndcg3) + 0.005)
      b_ap.append(float(ap))
      b_rr.append(float(rr) + 0.01)
  bu_min, bu_max = 0.775, 0.900
  bd_min, bd_max = 0.300, 0.425
  filename = 'tagrecom_yfcc10k_beta.eps'
  plot_tune('log $\\beta$', b_x, b_p3, b_f3, b_ndcg3, b_ap, b_rr, bu_min, bu_max, bd_min, bd_max, filename)


parser = argparse.ArgumentParser()
parser.add_argument('task', type=str, help='conv|tune')
args = parser.parse_args()

curmod = sys.modules[__name__]
func = getattr(curmod, args.task)
func()



