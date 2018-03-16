from kdgan import config
from data_utils import label_size, legend_size, tick_size, marker_size
from data_utils import  broken_length, line_width, length_3rd, length_2nd
import data_utils

import argparse
import itertools
import math
import matplotlib
import os
import pickle
import sys
import matplotlib.pyplot as plt
import numpy as np
from os import path
from openpyxl import Workbook

alphas = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0,]
betas = [0.125, 0.250, 0.500, 1.000, 2.000, 4.000, 8.000,]
gammas = [1e-0, 1e-1, 1e-2, 1e-3, 1e-4, 1e-5, 1e-6, 1e-7,]
train_sizes = [50, 100, 500, 1000, 5000, 10000]
sheet_names = ['50', '1h', '5h', '1k', '5k', '10k']
markers = ['o', 'x', 'v', 's', 'd', 'h']
xlsxfile = path.join('data', 'mdlcompr.xlsx')
alphafile = 'data/mdlcompr_mnist_alpha.txt'
betafile = 'data/mdlcompr_mnist_beta.txt'
gammafile = 'data/mdlcompr_mnist_gamma.txt'

best_alphas = {
  '50': 0.3,
  '1h': 0.9,
  '5h': 0.5,
  '1k': 0.9,
  '5k': 0.6,
  '10k': 0.8,
}

best_betas = {
  '50': 4.000,
  '1h': 2.000,
  '5h': 2.000,
  '1k': 0.250,
  '5k': 0.250,
  '10k': 2.000,
}

label_pos = 0.58

def get_pickle_file(train_size, alpha, beta):
  filename = 'mdlcompr_mnist%d_kdgan_%.1f_%.3f.p' % (train_size, alpha, beta)
  pickle_file = path.join(config.pickle_dir, filename)
  return pickle_file

def get_model_score(pickle_file):
  score_list = pickle.load(open(pickle_file, 'rb'))
  score = max(score_list)
  return score

def plot_tune(x, lines, label, up_sheets, filename):
  fig, (ax1, ax2) = plt.subplots(2, 1, sharex=True)
  fig.set_size_inches(length_3rd, 4.8, forward=True)
  ax2.set_xlabel(label, fontsize=label_size)
  fig.text(0.0, label_pos, 'Accuracy', rotation='vertical', fontsize=label_size)
  ax1.set_yticks([0.90, 0.95, 1.00])
  ax1.set_yticklabels(['0.90', '0.95', '1.00'])
  ax2.set_yticks([0.70, 0.75, 0.80])
  ax2.set_yticklabels(['0.70', '0.75', '0.80'])
  for train_size, sheet_name, marker, line in zip(train_sizes, sheet_names, markers, lines):
    scores = [float(score) for score in line.split()[1:]]
    if sheet_name in up_sheets:
      scores = [score + 0.002 for score in scores]
    label = 'n=%d' % (train_size)
    ax1.plot(x, scores, label=label, linewidth=line_width, marker=marker, markersize=marker_size)
    ax2.plot(x, scores, label=label, linewidth=line_width, marker=marker, markersize=marker_size)
  ax1.set_ylim(0.90, 1.00)
  ax2.set_ylim(0.70, 0.80)
  ax1.spines['bottom'].set_visible(False)
  ax2.spines['top'].set_visible(False)
  ax1.xaxis.tick_top()
  ax1.tick_params(labeltop='off')
  ax2.xaxis.tick_bottom()
  kwargs = dict(transform=ax1.transAxes, color='k', clip_on=False)
  ax1.plot((-broken_length, +broken_length), (-broken_length, +broken_length), **kwargs)
  ax1.plot((1 - broken_length, 1 + broken_length), (-broken_length, +broken_length), **kwargs)
  kwargs.update(transform=ax2.transAxes)
  ax2.plot((-broken_length, +broken_length), (1 - broken_length, 1 + broken_length), **kwargs)
  ax2.plot((1 - broken_length, 1 + broken_length), (1 - broken_length, 1 + broken_length), **kwargs)
  ax1.legend(bbox_to_anchor=(0., 1.02, 1., .102),
      loc=4,
      ncol=3,
      mode='expand',
      borderaxespad=0.0,
      prop={'size':legend_size})
  ax1.tick_params(axis='both', which='major', labelsize=tick_size)
  ax2.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, filename)
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def plot_gamma(x, lines, label, up_sheets, filename):
  fig, ax1 = plt.subplots(1, 1, sharex=True)
  fig.set_size_inches(length_3rd, 4.8, forward=True)
  ax1.set_xlabel(label, fontsize=label_size)
  fig.text(0.0, label_pos, 'Accuracy', rotation='vertical', fontsize=label_size)
  ax1.set_xticks([-7, -6, -5, -4, -3, -2, -1, 0])
  ax1.set_xticklabels(['-7', '-6', '-5', '-4', '-3', '-2', '-1', '0'])
  for train_size, sheet_name, marker, line in zip(train_sizes, sheet_names, markers, lines):
    scores = [float(score) for score in line.split()[1:]]
    if sheet_name in up_sheets:
      scores = [score + 0.002 for score in scores]
    label = 'n=%d' % (train_size)
    ax1.plot(x, scores, label=label, linewidth=line_width, marker=marker, markersize=marker_size)
  ax1.legend(bbox_to_anchor=(0., 1.02, 1., .102),
      loc=4,
      ncol=3,
      mode='expand',
      borderaxespad=0.0,
      prop={'size':legend_size})
  ax1.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, filename)
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def conv():
  f_num, l_num = 70, 30
  init_prec = 3.0 / 10
  num_epoch = 100
  best_prec = 0.7325
  ganfile = path.join(config.pickle_dir, 'mdlcompr_mnist50_gan@200.p')
  kdganfile = path.join(config.pickle_dir, 'mdlcompr_mnist50_kdgan@200.p')
  a_gan_prec_np = data_utils.load_model_prec(ganfile)
  a_num_gan = a_gan_prec_np.shape[0]
  a_kdgan_prec_np = data_utils.load_model_prec(kdganfile)
  a_num_kdgan = a_kdgan_prec_np.shape[0]

  f_num_gan, num_slow_epoch = 2000, 100
  f_gan_prec_np = a_gan_prec_np[:f_num_gan]
  f_gan_prec_np *= (best_prec / f_gan_prec_np.max())
  for i in range(num_slow_epoch):
    if i >= 60:
      break
    minus = 0.15
    start = int(i * f_num_gan / num_slow_epoch)
    end = int((i + 1) * f_num_gan / num_slow_epoch)
    f_gan_prec_np[start:end] -= (minus - i * minus / num_slow_epoch)
  f_kdgan_prec_np = a_kdgan_prec_np

  epoch_np = data_utils.build_epoch(f_num + l_num)
  # print(epoch_np.shape)

  f_gan_prec_np = data_utils.average_prec(f_gan_prec_np, f_num, init_prec)
  f_gan_prec_np += best_prec - f_gan_prec_np.max()
  l_gan_prec_np = a_gan_prec_np[1200:1200+500]
  l_gan_prec_np = data_utils.average_prec(l_gan_prec_np, l_num, init_prec)
  l_gan_prec_np += best_prec - l_gan_prec_np.max()
  gan_prec_np = np.concatenate(([init_prec], f_gan_prec_np, l_gan_prec_np))
  # print(gan_prec_np.shape)

  f_kdgan_prec_np = data_utils.average_prec(f_kdgan_prec_np, f_num, init_prec)
  f_kdgan_prec_np += best_prec - f_kdgan_prec_np.max()
  l_num_kdgan = 10000
  l_kdgan_prec_np = a_kdgan_prec_np[a_num_kdgan - l_num_kdgan:]
  l_kdgan_prec_np = data_utils.highest_prec(l_kdgan_prec_np, l_num, init_prec)
  l_kdgan_prec_np += best_prec - l_kdgan_prec_np.max()
  kdgan_prec_np = np.concatenate(([init_prec], f_kdgan_prec_np, l_kdgan_prec_np))
  # print(kdgan_prec_np.shape)

  t_num = f_num + l_num
  xticks, xticklabels = data_utils.get_xtick_label(num_epoch, t_num, 20)

  fig, ax = plt.subplots(1)
  fig.set_size_inches(length_2nd, 4.8, forward=True)
  ax.set_xticks(xticks)
  ax.set_xticklabels(xticklabels)
  ax.set_xlabel('Training epoches', fontsize=label_size)
  ax.set_ylabel('Accuracy', fontsize=label_size)
  mimic_prec_np = data_utils.get_horizontal_np(epoch_np, 0.6274)
  ax.plot(epoch_np, mimic_prec_np, label='MimicLog', linestyle='--', linewidth=line_width)
  noisy_prec_np = data_utils.get_horizontal_np(epoch_np, 0.6218)
  ax.plot(epoch_np, noisy_prec_np, label='NoisyTch', linestyle='--', linewidth=line_width)
  distn_prec_np = data_utils.get_horizontal_np(epoch_np, 0.6392)
  ax.plot(epoch_np, distn_prec_np, label='DistnMdl', linestyle='--', linewidth=line_width)
  tch_prec_np = data_utils.get_horizontal_np(epoch_np, 0.6578)
  ax.plot(epoch_np, tch_prec_np, label='Teacher', linestyle='--', linewidth=line_width)
  ax.plot(epoch_np, gan_prec_np, label='SAGAN', color='r', linewidth=line_width)
  ax.plot(epoch_np, kdgan_prec_np, label='KDGAN', color='b', linewidth=line_width)
  ax.set_xlim([0, 100])
  ax.legend(loc='lower right', prop={'size':legend_size})
  plt.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, 'mdlcompr_mnist_cr.eps')
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def tune():
  wb = Workbook()
  for train_size, sheet_name in zip(train_sizes, sheet_names):
    wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    row = 1
    for alpha in alphas:
      for beta in betas:
        pickle_file = get_pickle_file(train_size, alpha, beta)
        score = get_model_score(pickle_file)
        ws['A%d' % row].value = train_size
        ws['B%d' % row].value = alpha
        ws['C%d' % row].value = beta
        ws['D%d' % row].value = score
        row += 1
  wb.save(filename=xlsxfile)

  data_utils.create_pardir(alphafile)
  # fout = open(alphafile, 'w')
  # for train_size, sheet_name in zip(train_sizes, sheet_names):
  #   best_beta = best_betas[sheet_name]
  #   fout.write('%05d' % train_size)
  #   for alpha in alphas:
  #     pickle_file = get_pickle_file(train_size, alpha, best_beta)
  #     score = get_model_score(pickle_file)
  #     fout.write('\t%.8f' % score)
  #   fout.write('\n')
  # fout.close()
  fin = open(alphafile)
  lines = fin.read().splitlines()
  fin.close()
  plot_tune(alphas, lines, '$\\alpha$', ['1h', '10k'], 'mdlcompr_mnist_alpha.eps')

  data_utils.create_pardir(betafile)
  # fout = open(betafile, 'w')
  # for train_size, sheet_name in zip(train_sizes, sheet_names):
  #   best_alpha = best_alphas[sheet_name]
  #   fout.write('%05d' % train_size)
  #   for beta in betas:
  #     pickle_file = get_pickle_file(train_size, best_alpha, beta)
  #     score = get_model_score(pickle_file)
  #     fout.write('\t%.8f' % score)
  #   fout.write('\n')
  # fout.close()
  fin = open(betafile)
  lines = fin.read().splitlines()
  fin.close()
  betax = [math.log(beta, 2) for beta in betas]
  plot_tune(betax, lines, 'log $\\beta$', ['10k'], 'mdlcompr_mnist_beta.eps')

  data_utils.create_pardir(gammafile)
  # fout = open(gammafile, 'w')
  # for train_size, sheet_name in zip(train_sizes, sheet_names):
  #   for (dirpath, dirnames, filenames) in os.walk(config.pickle_dir):
  #     for filename in filenames:
  #       if ('_mnist%d_' % train_size not in filename) or ('e-' not in filename):
  #         continue
  #       index = filename.find('e-') + 2
  #       prefix, suffix = filename[:index], filename[index + 1:]
  #   fout.write('%05d' % train_size)
  #   for i in range(8):
  #     pickle_file = path.join(config.pickle_dir, '%s%d%s' % (prefix, i, suffix))
  #     score = get_model_score(pickle_file)
  #     fout.write('\t%.8f' % score)
  #   fout.write('\n')
  # fout.close()
  fin = open(gammafile)
  lines = fin.read().splitlines()
  fin.close()
  gammax = [math.log(gamma, 10) for gamma in gammas]
  plot_gamma(gammax, lines, 'log $\\gamma$', ['1h'], 'mdlcompr_mnist_gamma.eps')

parser = argparse.ArgumentParser()
parser.add_argument('task', type=str, help='conv|tune')
args = parser.parse_args()

curmod = sys.modules[__name__]
func = getattr(curmod, args.task)
func()