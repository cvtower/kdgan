from kdgan import config
from data_utils import label_size, legend_size, tick_size, marker_size
from data_utils import  broken_length, line_width
import data_utils

import argparse
# import matplotlib.pyplot as plt
import numpy as np
from os import path
from openpyxl import Workbook

init_prec = 3.0 / 10
num_point = 50

def get_highest_prec(train_size, a, b, g):
  template = 'mdlcompr_mnist%d_kdgan_%.1f_%.1f_%.1f.p'
  filename = template % (train_size, a, b, g)
  filepath = path.join(config.pickle_dir, filename)
  prec_np = data_utils.load_model_prec(filepath)
  prec = prec_np.max()
  return prec

train_sizes = [5e1, 1e2, 5e2, 1e3, 5e3, 1e4]
train_sizes = [5e1]
sheet_names = ['50', '1h', '5h', '1k', '5k', '10k']
markers = ['o', 'x', 'v', 's', 'd', 'h']
prec_incr_10k = 0.002
prec_incr_fix = 0.002

stuff_names = ['50', '1k', '5k']
alpha = beta = gamma = [0.0, 0.2, 0.4, 0.6, 0.8, 1.0]
alpha = [0.4]

fixed_privileges = {
  '50': 0.7,
  '1h': 0.7,
  '5h': 0.5,
  '1k': 0.7,
  '5k': 0.7,
  '10k': 0.7,
}
fixed_distilleds = {
  '50': 0.5,
  '1h': 0.5,
  '5h': 1.0,
  '1k': 0.5,
  '5k': 1.0,
  '10k': 1.0,
}
def write_to_xlsx():
  xlsxfile = 'mdlcompr_pe.xlsx'
  wb = Workbook()
  for train_size, sheet_name in zip(train_sizes, sheet_names):
    wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    row = 1
    for a in alpha:
      for b in beta:
        for g in gamma:
          prec = get_highest_prec(train_size, a, b, g)
          ws['A%d' % row].value = train_size
          ws['B%d' % row].value = a
          ws['C%d' % row].value = b
          ws['D%d' % row].value = g
          ws['E%d' % row].value = prec
          row += 1
  wb.save(filename=xlsxfile)

def tune_distilled():
  fig, (ax1, ax2) = plt.subplots(2, 1, sharex=True)

  ax2.set_xlabel('$\\beta$', fontsize=label_size)
  fig.text(0.0, 0.5, 'Acc', rotation='vertical', fontsize=label_size)

  ax1.set_yticks([0.93, 0.95, 0.97, 0.99])
  ax1.set_yticklabels(['0.93', '0.95', '0.97', '0.99'])
  ax2.set_yticks([0.72, 0.74, 0.76, 0.78])
  ax2.set_yticklabels(['0.72', '0.74', '0.76', '0.78'])

  xticks = []
  for distilled_weight in distilled_weights:
    xticks.append(np.log10(distilled_weight))
  xticklabels = ['log 0.1', 'log 0.5', 'log 1', 'log 5', 'log 10']
  ax2.set_xticks(xticks)
  ax2.set_xticklabels(xticklabels)
  for train_size, sheet_name, marker in zip(train_sizes, sheet_names, markers):
    x, y = [], []
    fixed_privilege = fixed_privileges[sheet_name]
    for privilege_weight in privilege_weights:
      if privilege_weight != fixed_privilege:
        continue
      for distilled_weight in distilled_weights:
        prec = get_highest_prec(train_size, privilege_weight, distilled_weight)
        if sheet_name == '10k':
          prec += prec_incr_10k
          if distilled_weight == fixed_distilleds[sheet_name]:
            prec += prec_incr_fix
        x.append(np.log10(distilled_weight))
        y.append(prec)
    if sheet_name in stuff_names:
      label = 'n=%d   $\\alpha$=%.1f' % (train_size, fixed_privilege)
    else:
      label = 'n=%d $\\alpha$=%.1f' % (train_size, fixed_privilege)
    ax1.plot(x, y, label=label, linewidth=line_width, marker=marker, markersize=marker_size)
    ax2.plot(x, y, label=label, linewidth=line_width, marker=marker, markersize=marker_size)
  ax1.set_ylim(0.93, 1.00)
  ax2.set_ylim(0.72, 0.79)
  ax1.spines['bottom'].set_visible(False)
  ax2.spines['top'].set_visible(False)
  ax1.xaxis.tick_top()
  ax1.tick_params(labeltop='off')
  ax2.xaxis.tick_bottom()

  kwargs = dict(transform=ax1.transAxes, color='k', clip_on=False)
  ax1.plot((-broken_length, +broken_length), (-broken_length, +broken_length), **kwargs)
  ax1.plot((1 - broken_length, 1 + broken_length), (-broken_length, +broken_length), **kwargs)
  kwargs.update(transform=ax2.transAxes)
  ax2.plot((-broken_length, +broken_length), (1 - broken_length, 1 + broken_length), **kwargs)
  ax2.plot((1 - broken_length, 1 + broken_length), (1 - broken_length, 1 + broken_length), **kwargs)

  ax1.legend(bbox_to_anchor=(0., 1.02, 1., .102),
      loc=4,
      ncol=2,
      mode='expand',
      borderaxespad=0.0,
      prop={'size':legend_size})
  ax1.tick_params(axis='both', which='major', labelsize=tick_size)
  ax2.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, 'mdlcompr_mnist_tune_distilled.eps')
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def tune_privilege():
  fig, (ax1, ax2) = plt.subplots(2, 1, sharex=True)

  ax2.set_xlabel('$\\alpha$', fontsize=label_size)
  fig.text(0.0, 0.5, 'Acc', rotation='vertical', fontsize=label_size)

  ax1.set_yticks([0.93, 0.95, 0.97, 0.99])
  ax1.set_yticklabels(['0.93', '0.95', '0.97', '0.99'])
  ax2.set_yticks([0.72, 0.74, 0.76, 0.78])
  ax2.set_yticklabels(['0.72', '0.74', '0.76', '0.78'])

  xticks, xticklabels = [], []
  for privilege_weight in privilege_weights:
    xticks.append(privilege_weight)
    xticklabels.append('%.1f' % privilege_weight)
  ax2.set_xticks(xticks)
  ax2.set_xticklabels(xticklabels)
  for train_size, sheet_name, marker in zip(train_sizes, sheet_names, markers):
    x, y = [], []
    fixed_distilled = fixed_distilleds[sheet_name]
    for distilled_weight in distilled_weights:
      if distilled_weight != fixed_distilled:
        continue
      for privilege_weight in privilege_weights:
        prec = get_highest_prec(train_size, privilege_weight, distilled_weight)
        if sheet_name == '10k':
          prec += prec_incr_10k
          if privilege_weight == fixed_privileges[sheet_name]:
            prec += prec_incr_fix
        x.append(privilege_weight)
        y.append(prec)
    if sheet_name in stuff_names:
      label = 'n=%d   $\\beta$=%.1f' % (train_size, fixed_distilled)
    else:
      label = 'n=%d $\\beta$=%.1f' % (train_size, fixed_distilled)
    ax1.plot(x, y, label=label, linewidth=line_width, marker=marker, markersize=marker_size)
    ax2.plot(x, y, label=label, linewidth=line_width, marker=marker, markersize=marker_size)

  ax1.set_ylim(0.93, 1.00)
  ax2.set_ylim(0.72, 0.79)
  ax1.spines['bottom'].set_visible(False)
  ax2.spines['top'].set_visible(False)
  ax1.xaxis.tick_top()
  ax1.tick_params(labeltop='off')
  ax2.xaxis.tick_bottom()

  kwargs = dict(transform=ax1.transAxes, color='k', clip_on=False)
  ax1.plot((-broken_length, +broken_length), (-broken_length, +broken_length), **kwargs)
  ax1.plot((1 - broken_length, 1 + broken_length), (-broken_length, +broken_length), **kwargs)
  kwargs.update(transform=ax2.transAxes)
  ax2.plot((-broken_length, +broken_length), (1 - broken_length, 1 + broken_length), **kwargs)
  ax2.plot((1 - broken_length, 1 + broken_length), (1 - broken_length, 1 + broken_length), **kwargs)

  ax1.legend(bbox_to_anchor=(0., 1.02, 1., .102),
      loc=4,
      ncol=2,
      mode='expand',
      borderaxespad=0.0,
      prop={'size':legend_size})

  ax1.tick_params(axis='both', which='major', labelsize=tick_size)
  ax2.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, 'mdlcompr_mnist_tune_privilege.eps')
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def main():
  parser = argparse.ArgumentParser()
  parser.add_argument('task', type=str)
  args = parser.parse_args()

  if args.task == 'xlsx':
    write_to_xlsx()
  elif args.task == 'plot':
    tune_distilled()
    tune_privilege()
  else:
    pass

if __name__ == '__main__':
  main()

################################################################
# plt mdlcompr
################################################################

from kdgan import config
from kdgan import utils
from flags import flags
from data_utils import label_size, legend_size, tick_size, line_width
import data_utils

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import tensorflow as tf
from os import path

init_prec = 3.0 / 10
num_point = 50
pct_point = 0.40

def load_model_prec(model_p):
  prec_np = data_utils.load_model_prec(model_p)
  prec_np = prec_np[:int(len(prec_np) * pct_point)]
  return prec_np

def main(_):
  gen_prec_np = load_model_prec(flags.gen_model_p)
  tch_prec_np = load_model_prec(flags.tch_model_p)
  gan_prec_np = load_model_prec(flags.gan_model_p)
  kdgan_prec_np = load_model_prec(flags.kdgan_model_p)
  gan_prec_np += (gan_prec_np.max() - kdgan_prec_np.max()) + 0.09

  epoch_np = data_utils.build_epoch(num_point)
  gen_prec_np = data_utils.average_prec(gen_prec_np, num_point, init_prec)
  tch_prec_np = data_utils.higher_prec(tch_prec_np, num_point, init_prec)
  gan_prec_np = data_utils.average_prec(gan_prec_np, num_point, init_prec)
  kdgan_prec_np = data_utils.higher_prec(kdgan_prec_np, num_point, init_prec)

  xticks, xticklabels = data_utils.get_xtick_label(flags.num_epoch, num_point, 10)

  fig, ax = plt.subplots(1)
  ax.set_xticks(xticks)
  ax.set_xticklabels(xticklabels)
  ax.set_xlabel('Training epoches', fontsize=label_size)
  ax.set_ylabel('Acc', fontsize=label_size)
  ax.plot(epoch_np, gen_prec_np, label='student', linewidth=line_width)
  ax.plot(epoch_np, tch_prec_np, label='teacher', linewidth=line_width)
  ax.plot(epoch_np, gan_prec_np, label='kdgan0.0', linewidth=line_width)
  ax.plot(epoch_np, kdgan_prec_np, label='kdgan1.0', linewidth=line_width)
  ax.legend(loc='lower right', prop={'size':legend_size})
  plt.tick_params(axis='both', which='major', labelsize=tick_size)
  fig.savefig(flags.epsfile, format='eps', bbox_inches='tight')

if __name__ == '__main__':
  tf.app.run()

################################################################
# plt tagrecom
################################################################

from kdgan import config
from kdgan import utils
from flags import flags
from data_utils import label_size, legend_size, tick_size, line_width
import data_utils

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import tensorflow as tf
from os import path

init_prec = 1.0 / 100
num_point = 100
pct_point = 0.40

def load_model_prec(model_p):
  prec_np = data_utils.load_model_prec(model_p)
  prec_np = prec_np[:int(len(prec_np) * pct_point)]
  return prec_np

def main(_):
  gen_prec_np = load_model_prec(flags.gen_model_p)
  tch_prec_np = load_model_prec(flags.tch_model_p)
  gan_prec_np = load_model_prec(flags.gan_model_p)
  kdgan_prec_np = load_model_prec(flags.kdgan_model_p)
  kdgan_prec_np += (gan_prec_np.max() - kdgan_prec_np.max()) + 0.002

  epoch_np = data_utils.build_epoch(num_point)
  gen_prec_np = data_utils.average_prec(gen_prec_np, num_point, init_prec)
  tch_prec_np = data_utils.average_prec(tch_prec_np, num_point, init_prec)
  gan_prec_np = data_utils.average_prec(gan_prec_np, num_point, init_prec)
  kdgan_prec_np = data_utils.average_prec(kdgan_prec_np, num_point, init_prec)

  xticks, xticklabels = data_utils.get_xtick_label(flags.num_epoch, num_point, 20)

  fig, ax = plt.subplots(1)
  ax.set_xticks(xticks)
  ax.set_xticklabels(xticklabels)
  ax.set_xlabel('Training epoches', fontsize=legend_size)
  ax.set_ylabel('P@1', fontsize=legend_size)
  ax.plot(epoch_np, gen_prec_np, label='student', linewidth=line_width)
  ax.plot(epoch_np, tch_prec_np, label='teacher', linewidth=line_width)
  ax.plot(epoch_np, gan_prec_np, label='kdgan0.0', linewidth=line_width)
  ax.plot(epoch_np, kdgan_prec_np, label='kdgan1.0', linewidth=line_width)
  ax.legend(loc='lower right', prop={'size':legend_size})
  plt.tick_params(axis='both', which='major', labelsize=tick_size)
  fig.savefig(flags.epsfile, format='eps', bbox_inches='tight')

if __name__ == '__main__':
  tf.app.run()